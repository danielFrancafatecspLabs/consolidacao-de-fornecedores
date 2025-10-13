from openpyxl import load_workbook
from typing import List, Dict
import unicodedata
import string


def _find_col_index(headers, names):
    for i, h in enumerate(headers):
        if h is None:
            continue
        hv = str(h).strip().lower()
        for name in names:
            if name.lower() == hv:
                return i
    return None


def parse_fornecedores_from_xlsx(path: str) -> List[Dict]:
    wb = load_workbook(filename=path, data_only=True)
    sheet_name = None
    # Try to find sheet with name similar to 'ANEXO 1 - DETALHES TÉCNICOS' or 'Detalhes Técnicos'
    for name in wb.sheetnames:
        lname = name.lower()
        if 'anexo' in lname and 'detal' in lname and 'técn' in lname:
            sheet_name = name
            break
    if not sheet_name:
        for name in wb.sheetnames:
            if 'detal' in name.lower() and 'técn' in name.lower():
                sheet_name = name
                break
    if not sheet_name:
        # fallback exact match
        if 'Detalhes Técnicos' in wb.sheetnames:
            sheet_name = 'Detalhes Técnicos'
    if not sheet_name:
        raise ValueError('Sheet "Detalhes Técnicos" not found')

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return []

    headers = [str(h).strip() if h is not None else None for h in rows[0]]
    # Normalize headers without enforcing required headers
    headers = [normalize_string(str(h)) if h is not None else None for h in rows[0]]

    # Replace empty headers with default names
    for i, header in enumerate(headers):
        if header is None or header.strip() == "":
            headers[i] = f"coluna_{i+1}"

    # tolerant mapping for required columns (including new ones)
    idx_fornecedor = _find_col_index(headers, ['Fornecedor', 'fornecedor'])
    idx_perfil = _find_col_index(headers, ['Perfil', 'perfil'])
    idx_hora = _find_col_index(headers, ['Hora', 'Horas', 'Horas Trabalhadas', 'horas'])
    idx_hh = _find_col_index(headers, ['H/H', 'H H', 'HH', 'h/h'])
    idx_qtde = _find_col_index(headers, ['Qtde de Recursos', 'Quantidade de Recursos', 'Qtde', 'Quantidade', 'qtde'])
    idx_aloc = _find_col_index(headers, ['Alocação (meses)', 'Alocação', 'Alocacao', 'Alocação meses', 'Alocacao (meses)'])
    idx_valor = _find_col_index(headers, ['Total', 'Valor total', 'Valor Total', 'Valor', 'Total (R$)'])

    print("Headers:", headers)
    print("First few rows:", rows[:5])

    data = {}
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue

        fornecedor = row[idx_fornecedor] if idx_fornecedor is not None else None
        perfil = row[idx_perfil] if idx_perfil is not None else None
        hora = row[idx_hora] if idx_hora is not None else None
        hh = row[idx_hh] if idx_hh is not None else None
        qtde = row[idx_qtde] if idx_qtde is not None else None
        aloc = row[idx_aloc] if idx_aloc is not None else None
        valor = row[idx_valor] if idx_valor is not None else None

        if fornecedor is None:
            continue
        nome = str(fornecedor).strip()

        # normalize numeric fields (handle comma decimal separators)
        def to_float(v):
            if v is None or v == '':
                return 0.0
            try:
                if isinstance(v, str):
                    v2 = v.replace('.', '').replace(',', '.')
                    return float(v2)
                return float(v)
            except Exception:
                return 0.0

        valor_f = to_float(valor)
        hora_f = to_float(hora) if hora is not None else None
        hh_f = to_float(hh) if hh is not None else None
        qtde_i = int(to_float(qtde)) if qtde is not None and to_float(qtde) != 0 else None
        aloc_i = int(to_float(aloc)) if aloc is not None and to_float(aloc) != 0 else None

        detalhe = {
            'perfil': perfil,
            'hora': hora_f,
            'hh': hh_f,
            'qtde_recursos': qtde_i,
            'alocacao_meses': aloc_i,
            'valor_total': valor_f,
        }

        if nome not in data:
            data[nome] = {
                'fornecedor': nome,
                'total': 0.0,
                'detalhes': []
            }

        data[nome]['total'] += valor_f
        data[nome]['detalhes'].append(detalhe)

    # Find the row containing the relevant headers
    header_row = None
    for i, row in enumerate(rows):
        if 'Fornecedor' in [str(cell).strip() for cell in row if cell]:
            header_row = i
            break

    if header_row is None:
        raise ValueError("Relevant headers not found in the sheet.")

    # Use the identified header row
    headers = [normalize_string(str(h)) if h is not None else None for h in rows[header_row]]

    # Skip rows above the header row
    rows = rows[header_row + 1:]

    # Skip irrelevant rows by finding the first row with meaningful data
    start_row = 0
    for i, row in enumerate(rows):
        if any(cell is not None for cell in row):
            start_row = i
            break

    rows = rows[start_row:]

    return list(data.values())

def normalize_string(value: str) -> str:
    """
    Normalize a string by removing accents, punctuation, and extra spaces.
    """
    if not value:
        return ""
    # Remove accents
    value = unicodedata.normalize('NFKD', value).encode('ASCII', 'ignore').decode('utf-8')
    # Remove punctuation
    value = value.translate(str.maketrans('', '', string.punctuation))
    # Convert to lowercase and strip extra spaces
    return value.strip().lower()

def validate_and_structure_data(data: List[Dict]) -> List[Dict]:
    """
    Valida, estrutura e normaliza os dados antes de inseri-los no banco de dados.
    """
    structured_data = []

    for item in data:
        # Validação de campos obrigatórios
        fornecedor = normalize_string(item.get('fornecedor'))
        if not fornecedor or not item.get('detalhes'):
            continue

        # Validação e estruturação dos detalhes
        detalhes_validos = []
        for detalhe in item['detalhes']:
            perfil = normalize_string(detalhe['perfil'])
            if perfil and detalhe['valor_total'] > 0:
                detalhe['perfil'] = perfil
                detalhes_validos.append(detalhe)

        if detalhes_validos:
            structured_data.append({
                'fornecedor': fornecedor,
                'total': sum(d['valor_total'] for d in detalhes_validos),
                'detalhes': detalhes_validos
            })

    return structured_data

# Atualize a função parse_fornecedores_from_xlsx para incluir validação e estruturação
def parse_and_validate_fornecedores(path: str) -> List[Dict]:
    """
    Processa, valida e estrutura os dados de um arquivo Excel.
    """
    raw_data = parse_fornecedores_from_xlsx(path)
    return validate_and_structure_data(raw_data)
